from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT_FILE = Path(__file__).with_name("Maintenance_Dashboard_Project.xlsx")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
ACCENT_FILL = PatternFill("solid", fgColor="D9EAF7")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=16, bold=True, color="1F1F1F")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


ROOM_INSPECTIONS = [
    ["2026-07-16", "North Tower", "1", "101", "Pass", "OK", "OK", "OK", "OK", "OK", "Ready for guest turnover", "Low", "WO-2401", "Chief Engineer", "2026-07-16"],
    ["2026-07-16", "North Tower", "1", "104", "Follow-Up", "Repair", "OK", "OK", "OK", "OK", "PTAC making noise overnight", "Medium", "WO-2402", "HVAC Tech", "2026-07-17"],
    ["2026-07-16", "South Tower", "2", "214", "Fail", "OK", "Repair", "OK", "Repair", "OK", "Bathroom leak and damaged vinyl edge", "High", "WO-2403", "Maintenance Supervisor", "2026-07-16"],
    ["2026-07-16", "South Tower", "3", "318", "Pass", "OK", "OK", "OK", "OK", "OK", "Deep-clean complete", "Low", "WO-2404", "Room Inspector", "2026-07-18"],
    ["2026-07-16", "Main Building", "Lobby", "Fitness", "Follow-Up", "OK", "OK", "Repair", "OK", "OK", "Treadmill outlet trips under load", "Medium", "WO-2405", "Electrician", "2026-07-18"],
]

INVENTORY_ITEMS = [
    ["INV-001", "HVAC", "PTAC Filter", "Guest Rooms", "Grainger", 12.5, 40, 20, 30, 4, "", "Yes", "Shelf A1", "Quarterly swap item"],
    ["INV-002", "Electrical", "LED 9W Bulb", "Guest Rooms", "Home Depot", 3.75, 80, 60, 120, 2, "", "Yes", "Shelf B2", "Warm white standard"],
    ["INV-003", "Plumbing", "1/2 in Supply Line", "Bathrooms", "Ferguson", 9.95, 12, 15, 20, 3, "", "No", "Shelf C1", "Common faucet repair"],
    ["INV-004", "Life Safety", "Smoke Detector", "Guest Rooms", "HD Supply", 28.0, 8, 10, 12, 5, "", "Yes", "Cage 1", "Serialized component"],
    ["INV-005", "General", "Door Closer", "Hallways", "Amazon Business", 45.0, 6, 4, 6, 2, "", "No", "Shelf D4", "ADA compliant"],
]

DAILY_TASKS = [
    ["D-001", "Boiler Room", "Check boiler pressure, temperature, and leak points", "AM Engineer", "Daily", "Every Day", 20, "Complete", "2026-07-16", "Document readings in log book"],
    ["D-002", "Pool", "Verify pool chemistry and equipment alarms", "PM Engineer", "Daily", "Every Day", 15, "In Progress", "2026-07-15", "Coordinate with front desk on closures"],
    ["D-003", "Public Areas", "Inspect lobby, corridors, and elevators for lighting or trip hazards", "Shift Engineer", "Daily", "Every Day", 25, "Not Started", "2026-07-15", "Open work order for deficiencies"],
    ["D-004", "Exterior", "Check parking lot, signage, and entry doors", "Shift Engineer", "Daily", "Every Day", 15, "Complete", "2026-07-16", "Note weather-related issues"],
]

WEEKLY_TASKS = [
    ["W-001", "Guest Rooms", "Test sample of PTAC units and clean return grilles", "HVAC Tech", "Weekly", "Monday", 90, "Not Started", "2026-07-09", "Rotate floors weekly"],
    ["W-002", "Fire Safety", "Test emergency lighting and exit signage", "Electrician", "Weekly", "Tuesday", 60, "Complete", "2026-07-15", "Log failures for replacement"],
    ["W-003", "Water Systems", "Flush low-use fixtures and inspect for odor", "Plumber", "Weekly", "Wednesday", 45, "In Progress", "2026-07-10", "Focus on out-of-order rooms"],
    ["W-004", "Exterior", "Inspect roof drains, downspouts, and grounds irrigation", "Grounds/Maintenance", "Weekly", "Friday", 75, "Not Started", "2026-07-11", "Escalate ponding issues immediately"],
]

MONTHLY_TASKS = [
    ["M-001", "Life Safety", "Inspect fire extinguishers and document seal/pressure status", "Chief Engineer", "Monthly", "1st Business Day", 75, "Not Started", "2026-06-30", "Coordinate with vendor if replacement is needed"],
    ["M-002", "Guest Rooms", "Deep inspection of 10% of rooms with furnishings checklist", "Maintenance Supervisor", "Monthly", "2nd Week", 180, "In Progress", "2026-06-20", "Use room inspection sheet for findings"],
    ["M-003", "Inventory", "Cycle count critical spares and update reorder points", "Storeroom Clerk", "Monthly", "3rd Week", 120, "Complete", "2026-07-02", "Match with purchasing records"],
    ["M-004", "Mechanical", "Lubricate motors, bearings, and door hardware", "Shift Engineer", "Monthly", "Last Week", 150, "Not Started", "2026-06-26", "Record used materials"],
]


def style_title(sheet, title: str, subtitle: str) -> None:
    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = subtitle
    sheet["A2"].alignment = Alignment(wrap_text=True)


def style_headers(sheet, row: int, columns: int) -> None:
    for cell in sheet[row][:columns]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def add_table(sheet, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def format_body(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def create_lists_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Lists")
    sheet.append(["Inspection Status", "System Status", "Priority", "Task Status", "Critical Item"])
    max_len = max(4, 4, 4, 4, 2)
    values = {
        "A": ["Pass", "Follow-Up", "Fail"],
        "B": ["OK", "Repair", "Replace"],
        "C": ["Low", "Medium", "High", "Critical"],
        "D": ["Not Started", "In Progress", "Complete", "Deferred"],
        "E": ["Yes", "No"],
    }
    for row in range(1, max_len + 1):
        sheet.cell(row=row + 1, column=1, value=values["A"][row - 1] if row <= len(values["A"]) else None)
        sheet.cell(row=row + 1, column=2, value=values["B"][row - 1] if row <= len(values["B"]) else None)
        sheet.cell(row=row + 1, column=3, value=values["C"][row - 1] if row <= len(values["C"]) else None)
        sheet.cell(row=row + 1, column=4, value=values["D"][row - 1] if row <= len(values["D"]) else None)
        sheet.cell(row=row + 1, column=5, value=values["E"][row - 1] if row <= len(values["E"]) else None)
    sheet.sheet_state = "hidden"


def create_dashboard_sheet(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Dashboard"
    style_title(
        sheet,
        "Fairfield Maintenance Command Center",
        "Interactive dashboard fed by the inspection, inventory, and recurring task sheets.",
    )
    sheet.freeze_panes = "A4"
    set_widths(sheet, {"A": 24, "B": 16, "C": 16, "D": 16, "E": 24, "F": 18, "G": 18})

    metrics = [
        ("B4", "Inspections Logged", "=COUNTA('Room Inspections'!A2:A250)"),
        ("B5", "Rooms Passed", '=COUNTIF(\'Room Inspections\'!E2:E250,"Pass")'),
        ("B6", "Needs Follow-Up", '=COUNTIF(\'Room Inspections\'!E2:E250,"Follow-Up")'),
        ("B7", "Rooms Failed", '=COUNTIF(\'Room Inspections\'!E2:E250,"Fail")'),
        ("E4", "Inventory Items", "=COUNTA('Parts Inventory'!A2:A250)"),
        ("E5", "Items to Reorder", '=COUNTIF(\'Parts Inventory\'!K2:K250,"Reorder")'),
        ("E6", "Inventory Value", "=SUMPRODUCT('Parts Inventory'!F2:F250,'Parts Inventory'!G2:G250)"),
        ("E7", "Critical Spares", '=COUNTIF(\'Parts Inventory\'!L2:L250,"Yes")'),
    ]
    for cell, label, formula in metrics:
        label_cell = sheet.cell(row=sheet[cell].row, column=sheet[cell].column - 1)
        label_cell.value = label
        label_cell.fill = ACCENT_FILL
        label_cell.font = Font(bold=True)
        label_cell.border = THIN_BORDER
        sheet[cell] = formula
        sheet[cell].border = THIN_BORDER
        sheet[cell].number_format = '$#,##0.00' if label == 'Inventory Value' else '0'

    sheet["A10"] = "Recurring Task Completion"
    sheet["A10"].font = Font(bold=True, size=12)
    task_headers = ["Cadence", "Completed", "Total", "Completion %"]
    for index, header in enumerate(task_headers, start=1):
        sheet.cell(row=11, column=index, value=header)
    style_headers(sheet, 11, len(task_headers))

    task_rows = [
        (12, "Daily", "Daily Tasks"),
        (13, "Weekly", "Weekly Tasks"),
        (14, "Monthly", "Monthly Tasks"),
    ]
    for row_num, label, source in task_rows:
        sheet.cell(row=row_num, column=1, value=label)
        sheet.cell(row=row_num, column=2, value=f'=COUNTIF(\'{source}\'!H2:H250,"Complete")')
        sheet.cell(row=row_num, column=3, value=f'=COUNTA(\'{source}\'!A2:A250)')
        sheet.cell(row=row_num, column=4, value=f'=IFERROR(B{row_num}/C{row_num},0)')
        sheet.cell(row=row_num, column=4).number_format = '0%'
    format_body(sheet, 12, 14, 1, 4)

    sheet["F10"] = "Follow-Up Snapshot"
    sheet["F10"].font = Font(bold=True, size=12)
    follow_up_headers = ["Room / Area", "Priority"]
    for index, header in enumerate(follow_up_headers, start=6):
        sheet.cell(row=11, column=index, value=header)
    style_headers(sheet, 11, 2)
    for idx, dashboard_row in enumerate(range(12, 17), start=2):
        sheet.cell(row=dashboard_row, column=6, value=f"=IF('Room Inspections'!E{idx}<>\"Pass\",'Room Inspections'!D{idx},\"\")")
        sheet.cell(row=dashboard_row, column=7, value=f"=IF(F{dashboard_row}=\"\",\"\",'Room Inspections'!L{idx})")
    format_body(sheet, 12, 16, 6, 7)

    bar_chart = BarChart()
    bar_chart.title = "Task Completion by Cadence"
    bar_chart.y_axis.title = "Tasks"
    bar_chart.x_axis.title = "Cadence"
    data = Reference(sheet, min_col=2, max_col=3, min_row=11, max_row=14)
    categories = Reference(sheet, min_col=1, min_row=12, max_row=14)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    bar_chart.height = 7
    bar_chart.width = 12
    sheet.add_chart(bar_chart, "A17")

    pie_chart = PieChart()
    pie_chart.title = "Inspection Status Mix"
    labels = Reference(sheet, min_col=1, min_row=4, max_row=7)
    data = Reference(sheet, min_col=2, min_row=4, max_row=7)
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.height = 7
    pie_chart.width = 10
    sheet.add_chart(pie_chart, "E17")


def create_room_inspection_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Room Inspections")
    style_title(sheet, "Detailed Room Inspections - Marriott Fairfield", "Track guest rooms and public spaces with pass/follow-up/fail outcomes and system-level notes.")
    headers = [
        "Date", "Wing/Area", "Floor", "Room/Space", "Inspection Status", "HVAC", "Plumbing", "Electrical", "Walls/Floors", "Life Safety", "Notes", "Priority", "Work Order", "Assigned To", "Target Date",
    ]
    sheet.append(headers)
    for row in ROOM_INSPECTIONS:
        sheet.append(row)
    style_headers(sheet, 3, len(headers))
    format_body(sheet, 4, 3 + len(ROOM_INSPECTIONS), 1, len(headers))
    sheet.freeze_panes = "A4"
    set_widths(sheet, {
        "A": 12, "B": 18, "C": 10, "D": 14, "E": 18, "F": 12, "G": 12, "H": 12, "I": 14, "J": 12, "K": 36, "L": 12, "M": 14, "N": 18, "O": 12,
    })
    add_table(sheet, f"A3:O{3 + len(ROOM_INSPECTIONS)}", "RoomInspectionTable")

    status_validation = DataValidation(type="list", formula1="=Lists!$A$2:$A$4", allow_blank=False)
    system_validation = DataValidation(type="list", formula1="=Lists!$B$2:$B$4", allow_blank=False)
    priority_validation = DataValidation(type="list", formula1="=Lists!$C$2:$C$5", allow_blank=False)
    for validation in (status_validation, system_validation, priority_validation):
        sheet.add_data_validation(validation)
    status_validation.add("E4:E250")
    system_validation.add("F4:J250")
    priority_validation.add("L4:L250")

    sheet.conditional_formatting.add("E4:E250", CellIsRule(operator="equal", formula=['"Pass"'], fill=GOOD_FILL))
    sheet.conditional_formatting.add("E4:E250", CellIsRule(operator="equal", formula=['"Follow-Up"'], fill=WARN_FILL))
    sheet.conditional_formatting.add("E4:E250", CellIsRule(operator="equal", formula=['"Fail"'], fill=BAD_FILL))
    sheet.conditional_formatting.add("L4:L250", CellIsRule(operator="equal", formula=['"Critical"'], fill=BAD_FILL))
    sheet.auto_filter.ref = f"A3:O{3 + len(ROOM_INSPECTIONS)}"


def create_inventory_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Parts Inventory")
    style_title(sheet, "Parts Inventory", "Manage critical spares, reorder triggers, vendors, and storage locations.")
    headers = [
        "Part ID", "Category", "Part Name", "Equipment/Area", "Vendor", "Unit Cost", "On Hand", "Min Level", "Reorder Qty", "Lead Time (Days)", "Reorder Status", "Critical Item", "Storage Location", "Notes",
    ]
    sheet.append(headers)
    for row_index, row in enumerate(INVENTORY_ITEMS, start=4):
        sheet.append(row)
        sheet[f"K{row_index}"] = f'=IF(G{row_index}<=H{row_index},"Reorder","OK")'
    style_headers(sheet, 3, len(headers))
    format_body(sheet, 4, 3 + len(INVENTORY_ITEMS), 1, len(headers))
    sheet.freeze_panes = "A4"
    set_widths(sheet, {
        "A": 12, "B": 14, "C": 18, "D": 18, "E": 18, "F": 12, "G": 10, "H": 10, "I": 12, "J": 14, "K": 14, "L": 12, "M": 16, "N": 28,
    })
    add_table(sheet, f"A3:N{3 + len(INVENTORY_ITEMS)}", "InventoryTable")

    critical_validation = DataValidation(type="list", formula1="=Lists!$E$2:$E$3", allow_blank=False)
    sheet.add_data_validation(critical_validation)
    critical_validation.add("L4:L250")

    sheet.conditional_formatting.add("K4:K250", CellIsRule(operator="equal", formula=['"Reorder"'], fill=WARN_FILL))
    sheet.conditional_formatting.add("L4:L250", CellIsRule(operator="equal", formula=['"Yes"'], fill=ACCENT_FILL))
    sheet.auto_filter.ref = f"A3:N{3 + len(INVENTORY_ITEMS)}"


def create_task_sheet(workbook: Workbook, title: str, subtitle: str, name: str, rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(name)
    style_title(sheet, title, subtitle)
    headers = [
        "Task ID", "Area", "Task", "Assigned To", "Frequency", "Due Day/Date", "Est. Minutes", "Status", "Last Completed", "Notes",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    style_headers(sheet, 3, len(headers))
    format_body(sheet, 4, 3 + len(rows), 1, len(headers))
    sheet.freeze_panes = "A4"
    set_widths(sheet, {
        "A": 12, "B": 16, "C": 42, "D": 18, "E": 12, "F": 16, "G": 12, "H": 14, "I": 14, "J": 28,
    })
    table_name = name.replace(" ", "") + "Table"
    add_table(sheet, f"A3:J{3 + len(rows)}", table_name)

    status_validation = DataValidation(type="list", formula1="=Lists!$D$2:$D$5", allow_blank=False)
    sheet.add_data_validation(status_validation)
    status_validation.add("H4:H250")
    sheet.conditional_formatting.add("H4:H250", CellIsRule(operator="equal", formula=['"Complete"'], fill=GOOD_FILL))
    sheet.conditional_formatting.add("H4:H250", CellIsRule(operator="equal", formula=['"In Progress"'], fill=WARN_FILL))
    sheet.conditional_formatting.add("H4:H250", CellIsRule(operator="equal", formula=['"Deferred"'], fill=BAD_FILL))
    sheet.auto_filter.ref = f"A3:J{3 + len(rows)}"


def build_workbook() -> Workbook:
    workbook = Workbook()
    create_dashboard_sheet(workbook)
    create_room_inspection_sheet(workbook)
    create_inventory_sheet(workbook)
    create_task_sheet(
        workbook,
        "Daily Maintenance Tasks",
        "Use this sheet to manage daily building checks, hotel operations walk-throughs, and shift handoffs.",
        "Daily Tasks",
        DAILY_TASKS,
    )
    create_task_sheet(
        workbook,
        "Weekly Maintenance Tasks",
        "Use this sheet to plan recurring weekly preventive maintenance work.",
        "Weekly Tasks",
        WEEKLY_TASKS,
    )
    create_task_sheet(
        workbook,
        "Monthly Maintenance Tasks",
        "Use this sheet to track higher-touch preventive maintenance and inventory routines.",
        "Monthly Tasks",
        MONTHLY_TASKS,
    )
    create_lists_sheet(workbook)
    return workbook


def main() -> None:
    workbook = build_workbook()
    workbook.save(OUTPUT_FILE)
    print(f"Created {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
